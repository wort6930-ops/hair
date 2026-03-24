import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
import os
from pathlib import Path
import re
from config import TEMP_DIR, EXCEL_DATE_STR

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generate Excel files with scraped data"""
    def __init__(self):
        Path(TEMP_DIR).mkdir(parents=True, exist_ok=True)
    def _extract_price_value(self, price_str: str) -> float:
        if not price_str or price_str == 'N/A':
            return 0.0
        match = re.search(r'[\d.]+', str(price_str))
        if match:
            try:
                return float(match.group())
            except ValueError:
                return 0.0
        return 0.0
    def create_category_workbook(self, category_name: str, subcategories_data: Dict[str, List[Dict]]) -> str:
        logger.info(f"Creating Excel workbook for category: {category_name}")
        wb = Workbook()
        wb.remove(wb.active)
        for subcategory_name, products in subcategories_data.items():
            ws = wb.create_sheet(title=self._sanitize_sheet_name(subcategory_name))
            self._populate_sheet(ws, products)
        summary_ws = wb.create_sheet(title="Summary", index=0)
        self._create_summary_sheet(summary_ws, subcategories_data)
        excel_dir = os.path.join(TEMP_DIR, "excel_files")
        Path(excel_dir).mkdir(parents=True, exist_ok=True)
        filename = f"{excel_dir}/{category_name}_{EXCEL_DATE_STR}.xlsx"
        wb.save(filename)
        logger.info(f"Excel file created: {filename}")
        return filename
    def _populate_sheet(self, ws, products: List[Dict]):
        headers = [
            'Product Name', 'Brand', 'Price', 'Old Price', 'Discount', 'SKU', 'Description', 'Rating', 'Reviews', 'Colors', 'Product URL', 'S3 Image Path', 'Image URL'
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1).value = product.get('name', '')
            ws.cell(row=row_num, column=2).value = product.get('brand', '')
            ws.cell(row=row_num, column=3).value = product.get('price', '')
            ws.cell(row=row_num, column=4).value = product.get('old_price', '')
            ws.cell(row=row_num, column=5).value = product.get('discount', '')
            ws.cell(row=row_num, column=6).value = product.get('sku', '')
            ws.cell(row=row_num, column=7).value = product.get('description', '')
            ws.cell(row=row_num, column=8).value = product.get('rating', '')
            ws.cell(row=row_num, column=9).value = product.get('reviews', '')
            ws.cell(row=row_num, column=10).value = product.get('colors', '')
            ws.cell(row=row_num, column=11).value = product.get('product_url', '')
            ws.cell(row=row_num, column=12).value = product.get('s3_image_path', '')
            ws.cell(row=row_num, column=13).value = product.get('image_url', '')
            for col_num in range(1, 14):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        column_widths = [30, 15, 12, 12, 10, 15, 40, 10, 15, 20, 30, 30, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        ws.row_dimensions[1].height = 30
        logger.info(f"Sheet populated with {len(products)} products")
    def _create_summary_sheet(self, ws, subcategories_data: Dict[str, List[Dict]]):
        headers = ['Subcategory', 'Product Count', 'Brands', 'Average Price']
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        row_num = 2
        for subcategory_name, products in subcategories_data.items():
            ws.cell(row=row_num, column=1).value = subcategory_name
            ws.cell(row=row_num, column=2).value = len(products)
            brands = set(product.get('brand', '') for product in products if product.get('brand'))
            ws.cell(row=row_num, column=3).value = ', '.join(brands) if brands else 'N/A'
            prices = [self._extract_price_value(product.get('price')) for product in products]
            avg_price = sum(prices) / len(prices) if prices else 0
            ws.cell(row=row_num, column=4).value = f"{avg_price:.2f}"
            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            row_num += 1
        total_products = sum(len(products) for products in subcategories_data.values())
        ws.cell(row=row_num, column=1).value = "TOTAL"
        ws.cell(row=row_num, column=2).value = total_products
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        for col_num in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        logger.info("Summary sheet created")
    def _sanitize_sheet_name(self, name: str) -> str:
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name[:31]
